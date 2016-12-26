from lxml import etree
from bs4 import BeautifulSoup
from openpyxl import Workbook

import requests
import random
import json
import datetime
import re

DAYS_PER_WEEK = 7

def get_courses_list(courses_count=20):
    coursera_url = 'https://www.coursera.org/sitemap~www~courses.xml'
    response = requests.get(coursera_url)
    root = etree.fromstring(response.content)

    return random.sample(
        [child[0].text for child in root],
        courses_count
    )


def get_course_info(course_url):
    course_data = {}
    response = requests.get(course_url)
    html_doc = BeautifulSoup(response.content, 'html.parser')
    course_data['course_url'] = course_url
    course_data['title'] = get_title(html_doc)
    course_data['language'] = get_language(html_doc)
    starts, commitment = get_start_date_and_commitment(html_doc, course_url)
    course_data['starts'] = starts
    course_data['commitment'] = commitment
    course_data['rating'] = get_rating(html_doc)
    return course_data


def get_title(html_doc):
    result_tag = html_doc.find(class_='title display-3-text')
    title = None
    if result_tag:
        title = result_tag.string
    return title


def get_language(html_doc):
    result_tag = html_doc.find(class_='language-info')
    language = None
    if result_tag and result_tag.contents:
        language = result_tag.contents[1]
    return language


def get_start_date_and_commitment(html_doc, course_url):
    starts = None
    commitment = 0
    course_info = html_doc.find('script', type='application/ld+json')
    if course_info:
        course = json.loads(course_info.string)['hasCourseInstance'][0]
        start_date = course.get('startDate')
        if start_date:
            start_date = datetime.datetime.strptime(
                course['startDate'],
                '%Y-%m-%d'
            )
            end_date = course.get('endDate')
            if end_date:    
                end_date = datetime.datetime.strptime(
                    course['endDate'],
                    '%Y-%m-%d'
                )
                delta = end_date - start_date
                commitment = delta.days / DAYS_PER_WEEK
            starts = start_date
    else:
        starts = get_start_date_from_api(course_url)

    return starts, commitment        


def get_start_date_from_api(course_url):
    starts = None
    api_url = 'https://api.coursera.org/api/courses.v1'
    match = re.search(r'/[^/]+$', course_url)
    slug = match.group(0)[1:]
    params = {
        'q': 'slug',
        'slug': slug,
        'fields': 'startDate'
    }
    response = requests.get(api_url, params=params)
    course = response.json()['elements'][0]
    start_date = course.get('startDate')
    if start_date:
        starts = datetime.datetime.fromtimestamp(
            convert_to_seconds(int(start_date))
        )
    return starts


def convert_to_seconds(milliseconds_timestamp):
    milliseconds_in_second = 1000
    return milliseconds_timestamp / milliseconds_in_second


def get_rating(html_doc):
    result_tag = html_doc.find(class_='ratings-text bt3-visible-xs')
    str_rating = ''
    if result_tag:
        str_rating = result_tag.string
    return float(str_rating[:4]) if str_rating else 0


def output_courses_info_to_xlsx(courses, filepath):
    wb = Workbook()
    ws = wb.active
    keys = ['title', 'starts', 'language',
        'commitment', 'rating', 'course_url'
    ]
    for column, key in enumerate(keys, start=1):
        ws.cell(row=1, column=column, value=key)

    for row, item in enumerate(courses, start=2):
        for column, key in enumerate(keys, start=1):
            ws.cell(row=row, column=column, value=item[key])

    wb.save(filepath)


if __name__ == '__main__':
    output_courses_info_to_xlsx(
        (get_course_info(course_url) for course_url in get_courses_list()),
        input('Enter filepath (.xlsx):')
    )
        